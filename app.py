import io
import re
import zipfile
import unicodedata
from datetime import datetime, date, time, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st
import openpyxl
from copy import copy as pycopy

# ✅ Bordures + gris sur Total
from openpyxl.styles import PatternFill, Border, Side

# =========================
# CONFIG (2 modes)
# =========================
APP_DIR = Path(__file__).resolve().parent

TEMPLATE_IMPERIUM_PATH = APP_DIR / "TEMPLATE_SUIVI_FINAL.xlsx"
TEMPLATE_YUMI_PATH = APP_DIR / "SUIVI GATO.xlsx"

HEADER_ROW = 9
DATA_START_ROW = 10

DECALAGE_MINUTES = 45
TV_DAY_CUTOFF_HOUR = 6  # journée TV: 00:00-05:59 => fin journée => +1440

# ✅ Règle validée: ordre + swap intelligent (sans anticipé) (fallback only)
SWAP_NEAR_MINUTES = 5
SWAP_PM1_FAR_MINUTES = 20
ANTI_VOL_NEXT_MAX = 20
LOOKAHEAD_STEPS = 1

# ✅ Compensation "proche" d'un backlog Non diffusé (J-1 etc.)
COMPENSATION_MAX_MINUTES = 20

# ✅ YUMI: on ne génère QUE ces chaînes (même si PM contient d'autres)
ALLOWED_YUMI = {"2M", "ALAOULA"}

# ✅ Matching optimal (corrige le cas EXEED: laisser un PM "Non diffusé" plutôt que forcer l'ordre)
USE_OPTIMAL_MATCHING = True
PM_SKIP_PENALTY = 30       # coût de laisser un PM en Non diffusé
REAL_SKIP_PENALTY = 30     # coût de laisser une diffusion sans PM (passage supp/comp)
HARD_MAX_MATCH = 180       # au-delà (minutes), on décourage fortement le match

FINAL_COLUMNS_IMPERIUM = [
    "datep",
    "supportp",
    "heure de diffusion",
    "Code PM",
    "Commentaire",
    "Message",
    "Produit",
    "Marque",
    "RaisonSociale",
    "FormatSec",
    "Avant",
    "Apres",
    "rangE",
    "encombE",
]

FINAL_COLUMNS_YUMI = [
    "Date",
    "Chaîne",
    "N° Mois",
    "Année",
    "Annonceur",
    "Marque",
    "Produit",
    "H.Début",
    "H.Fin",
    "Durée",
    "Code Ecran",
    "Code Ecran PM",
    "Commentaire",
    "Programme après",
    "Programme avant",
    "Position",
    "Rang",
    "Encombrement",
    "Storyboard",
    "TM%",
    "TME",
]

# =========================
# UI config
# =========================
st.set_page_config(page_title="Suivi Pige — Automatisation (PM 2026 unique)", page_icon="📊", layout="wide")

st.markdown("""
<style>
.main { background-color: #f8fafc; }
.stButton>button {
    width: 100%;
    border-radius: 8px;
    height: 3.5em;
    background-color: #7289DA;
    color: white;
    font-weight: bold;
    border: none;
}
.stDownloadButton>button {
    width: 100%;
    border-radius: 8px;
    background-color: #43b581 !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)

# =========================
# STATE
# =========================
if "client_files" not in st.session_state:
    st.session_state.client_files = None
if "zip_bytes" not in st.session_state:
    st.session_state.zip_bytes = None
if "last_run_info" not in st.session_state:
    st.session_state.last_run_info = None

# =========================
# Utils
# =========================
def norm_txt(x):
    if x is None:
        return ""
    s = str(x).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s

def normalize_brand(name: str) -> str:
    s = norm_txt(name)
    s = re.sub(r"\bPM\b", " ", s)
    s = re.sub(r"\bRAMADAN\b", " ", s)
    s = re.sub(r"\bTV\b|\bRADIO\b|\bOOH\b", " ", s)
    s = re.sub(r"\bV\d+\b", " ", s)
    s = re.sub(r"\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b", " ", s)
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def brand_key(name: str) -> str:
    return normalize_brand(name).replace(" ", "")

def normalize_support(sup: str) -> str:
    s = norm_txt(sup)
    s = s.replace(" ", "")
    s = re.sub(r"[^A-Z0-9]+", "", s)
    s = s.replace("TV", "")
    if s in ("AOULA", "ALAOUOLA", "ALAOULA"):
        return "ALAOULA"
    return s

def safe_sheet_name(s: str) -> str:
    s = re.sub(r"[:\\/*?\[\]]", " ", str(s))
    s = re.sub(r"\s+", " ", s).strip()
    return s[:31] if len(s) > 31 else s

def find_column(df: pd.DataFrame, candidates: list[str]):
    cols = {c: norm_txt(c) for c in df.columns}
    for c, cn in cols.items():
        for cand in candidates:
            if norm_txt(cand) in cn:
                return c
    return None

def extract_first_time_from_text(val):
    """Ex: '21:12 | 21:22' -> time(21,12)"""
    if val is None:
        return None
    s = str(val)
    found = re.findall(r"(\d{1,2})\s*[:hH]\s*(\d{2})", s)
    if not found:
        return None
    times = []
    for hh, mm in found:
        hh_i = int(hh)
        mm_i = int(mm)
        if 0 <= hh_i <= 29 and 0 <= mm_i <= 59:
            times.append((hh_i, mm_i))
    if not times:
        return None
    hh_i, mm_i = min(times)  # plus tôt
    return time(hh_i % 24, mm_i, 0)

def to_excel_time(val):
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except:
        pass

    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return None
        return val.to_pydatetime().time()

    if isinstance(val, (float, int)):
        seconds = int(round(float(val) * 86400))
        seconds = max(0, min(seconds, 86399))
        return time(seconds // 3600, (seconds % 3600) // 60, seconds % 60)

    if isinstance(val, str):
        t0 = extract_first_time_from_text(val)
        if t0 is not None:
            return t0

    try:
        s = str(val).strip().replace("h", ":").replace("H", ":")
        t = pd.to_datetime(s, errors="coerce")
        if pd.isna(t):
            return None
        return t.to_pydatetime().time()
    except:
        return None

def code_hhmm_digits(x):
    """
    Convertit '24-15', '24:15', '2415R', '24h15' -> '2415'
    Ignore suffix letters.
    """
    if x is None:
        return None
    s = str(x).strip().upper().replace("H", ":")
    m = re.search(r"(\d{3,4})", s)
    if m:
        return m.group(1)
    m2 = re.search(r"(\d{1,2})\s*[:\-]\s*(\d{2})", s)
    if m2:
        hh = int(m2.group(1))
        mm = int(m2.group(2))
        return f"{hh:02d}{mm:02d}"
    return None

def parse_codepm_time(code_pm: str):
    """
    Supporte: 1600R / 1600A / 16:00 / 16-00 / 24-15 etc.
    Retourne: (match_time, overnight_bool, tv_minutes_raw)
    tv_minutes_raw = hh*60+mm (hh peut être >=24)
    """
    hhmm = code_hhmm_digits(code_pm)
    if not hhmm:
        return None, False, None

    if len(hhmm) == 3:
        hh = int(hhmm[0]); mm = int(hhmm[1:])
    else:
        hh = int(hhmm[:2]); mm = int(hhmm[2:])

    if mm < 0 or mm > 59:
        return None, False, None

    tv_minutes = hh * 60 + mm
    overnight = hh >= 24
    hh_mod = hh % 24
    return time(hh_mod, mm, 0), overnight, tv_minutes

def real_tv_minutes(t: time, cutoff_hour: int = TV_DAY_CUTOFF_HOUR):
    if t is None:
        return None
    m = t.hour * 60 + t.minute
    if t.hour < cutoff_hour:
        m += 1440
    return m

def pm_tv_minutes_tvday(code_pm: str, fallback=None, cutoff_hour: int = TV_DAY_CUTOFF_HOUR):
    """
    Minutes TV-DAY pour PM:
    - si hh < cutoff (et hh < 24) => +1440
    - si hh >= 24 => garde tel quel
    """
    try:
        if fallback is not None and pd.notna(fallback):
            tvm = int(fallback)
        else:
            tvm = None
    except:
        tvm = None

    if tvm is None:
        _, _, tvm = parse_codepm_time(code_pm)

    if tvm is None:
        return 10**9

    hh = tvm // 60
    if hh < 24 and hh < cutoff_hour:
        tvm += 1440
    return tvm

def make_zip(files: dict[str, bytes]) -> bytes:
    bio = io.BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in files.items():
            z.writestr(name, data)
    return bio.getvalue()

def get_min_max_date_from_raw(df_in: pd.DataFrame, mode: str):
    col = find_column(df_in, ["datep", "date"]) if mode == "Suivi Imperium" else find_column(df_in, ["date"])
    if not col:
        return None, None
    s = pd.to_datetime(df_in[col], errors="coerce").dropna()
    if s.empty:
        return None, None
    return s.dt.date.min(), s.dt.date.max()

# =========================
# Swap decision (fallback)
# =========================
def should_swap(rt_i, pm_i, pm_j, rt_next=None):
    if rt_i is None:
        return False

    d1 = abs(rt_i - pm_i) if pm_i < 10**9 else 10**9
    d2 = abs(rt_i - pm_j) if pm_j < 10**9 else 10**9

    if not (d2 <= SWAP_NEAR_MINUTES and d1 > SWAP_PM1_FAR_MINUTES):
        return False

    if rt_next is None:
        return True

    e1 = abs(rt_next - pm_i) if pm_i < 10**9 else 10**9
    e2 = abs(rt_next - pm_j) if pm_j < 10**9 else 10**9

    if e2 <= ANTI_VOL_NEXT_MAX and e2 < e1 and e1 > d1:
        return False

    return True

# =========================
# Matching engine:
# - Exact lock (digits)
# - Optimal DP matching (default)
# - Fallback ordre+swap if USE_OPTIMAL_MATCHING=False
# =========================
def match_day_exact_then_order_swap(rt_minutes, pm_minutes, real_codes, pm_codes):
    n = len(rt_minutes)
    m = len(pm_minutes)

    assign = [None] * n
    used_pm = set()
    locked_real = set()

    # 0) exact match first (digits), choose first available PM
    if real_codes and pm_codes and len(real_codes) == n and len(pm_codes) == m:
        for i in range(n):
            rc = real_codes[i]
            if not rc:
                continue
            for j in range(m):
                if j in used_pm:
                    continue
                if pm_codes[j] == rc:
                    assign[i] = j
                    used_pm.add(j)
                    locked_real.add(i)
                    break

    real_rem = [i for i in range(n) if i not in locked_real]
    pm_rem = [j for j in range(m) if j not in used_pm]

    if not USE_OPTIMAL_MATCHING:
        # order + swap fallback (without breaking exact match)
        remaining_real = [i for i in range(n) if assign[i] is None]
        k = min(len(remaining_real), len(pm_rem))
        for idx in range(k):
            assign[remaining_real[idx]] = pm_rem[idx]

        for i in range(n - 1):
            if i in locked_real or (i + 1) in locked_real:
                continue
            j1 = assign[i]
            j2 = assign[i + 1]
            if j1 is None or j2 is None:
                continue
            if j2 != j1 + 1:
                continue
            rt_i = rt_minutes[i]
            rt_next = rt_minutes[i + 1]
            if should_swap(rt_i, pm_minutes[j1], pm_minutes[j2], rt_next=rt_next):
                assign[i], assign[i + 1] = assign[i + 1], assign[i]
        return assign

    # DP alignment
    INF = 10**15
    A = len(real_rem)
    B = len(pm_rem)

    dp = [[INF] * (B + 1) for _ in range(A + 1)]
    prev = [[None] * (B + 1) for _ in range(A + 1)]
    dp[0][0] = 0

    for b in range(1, B + 1):
        dp[0][b] = dp[0][b - 1] + PM_SKIP_PENALTY
        prev[0][b] = ("SKIP_PM", 0, b - 1)

    for a in range(1, A + 1):
        dp[a][0] = dp[a - 1][0] + REAL_SKIP_PENALTY
        prev[a][0] = ("SKIP_REAL", a - 1, 0)

    def match_cost(rt, pmv):
        if rt is None or pmv is None or rt >= 10**9 or pmv >= 10**9:
            return HARD_MAX_MATCH * 10
        d = abs(rt - pmv)
        if d > HARD_MAX_MATCH:
            return d + 5000
        return d

    for a in range(1, A + 1):
        i = real_rem[a - 1]
        rt = rt_minutes[i]
        for b in range(1, B + 1):
            j = pm_rem[b - 1]
            pmv = pm_minutes[j]

            best = dp[a - 1][b - 1] + match_cost(rt, pmv)
            best_prev = ("MATCH", a - 1, b - 1)

            c_spm = dp[a][b - 1] + PM_SKIP_PENALTY
            if c_spm < best:
                best = c_spm
                best_prev = ("SKIP_PM", a, b - 1)

            c_sreal = dp[a - 1][b] + REAL_SKIP_PENALTY
            if c_sreal < best:
                best = c_sreal
                best_prev = ("SKIP_REAL", a - 1, b)

            dp[a][b] = best
            prev[a][b] = best_prev

    a, b = A, B
    pairs = []
    while a > 0 or b > 0:
        step, pa, pb = prev[a][b]
        if step == "MATCH":
            i = real_rem[a - 1]
            j = pm_rem[b - 1]
            pairs.append((i, j))
        a, b = pa, pb

    for i, j in pairs:
        assign[i] = j

    return assign

# =========================
# Template loader
# =========================
def load_template_workbook(mode: str) -> openpyxl.Workbook:
    if mode == "Suivi YUMI":
        return openpyxl.load_workbook(TEMPLATE_YUMI_PATH)
    return openpyxl.load_workbook(TEMPLATE_IMPERIUM_PATH)

# =========================
# PM 2026 parsing
# =========================
def parse_sheet_name(sheet_name: str, known_support_norms: set[str] | None = None):
    tokens = sheet_name.strip().split()
    if not tokens:
        return None, None, None

    vocab = set(known_support_norms or set())
    vocab |= {"2M", "MBC5", "ALAOULA"}

    best = None
    for k in range(1, min(4, len(tokens)) + 1):
        sup = " ".join(tokens[-k:])
        sup_norm = normalize_support(sup)
        if sup_norm in vocab:
            brand = " ".join(tokens[:-k]).strip()
            if brand:
                best = (brand, sup, sup_norm)
    if best:
        return best

    sup = tokens[-1]
    brand = " ".join(tokens[:-1]).strip() or sheet_name
    return brand, sup, normalize_support(sup)

def read_pm_2026_workbook(pm_bytes: bytes, known_support_norms: set[str]) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(pm_bytes), data_only=True)
    recs = []
    for sh in wb.sheetnames:
        ws = wb[sh]
        brand, sup_disp, sup_norm = parse_sheet_name(sh, known_support_norms)
        if not brand:
            continue

        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, 1).value
            code = ws.cell(r, 2).value

            if d is None and code is None:
                continue
            if code is None or str(code).strip() == "":
                continue

            d_parsed = pd.to_datetime(d, errors="coerce", dayfirst=True)
            if pd.isna(d_parsed):
                continue
            d_date = d_parsed.date()

            codepm = str(code).strip()
            _, overnight, tvm_raw = parse_codepm_time(codepm)

            recs.append({
                "PM_FILE_BRAND": brand,
                "PM_FILE_BRAND_N": brand_key(brand),
                "Date": pd.to_datetime(d_date),
                "date_only": d_date,
                "supportp": sup_disp,
                "support_norm": sup_norm,
                "Code PM": codepm,
                "Overnight": overnight,
                "PM_TV_MIN": tvm_raw,
            })

    return pd.DataFrame(recs)

# =========================
# DATA builders
# =========================
def build_final_df_from_imperium(df_imp: pd.DataFrame, date_min: date, date_max: date) -> pd.DataFrame:
    col_date = find_column(df_imp, ["datep", "date"])
    col_sup  = find_column(df_imp, ["supportp", "support", "chaine", "station"])
    col_time = find_column(df_imp, ["heurep", "heure"])
    col_mar  = find_column(df_imp, ["marque"])
    if not all([col_date, col_sup, col_time, col_mar]):
        raise ValueError("DATA IMPERIUM: colonnes minimales manquantes (datep/supportp/heurep/marque).")

    df = df_imp.copy()
    df["datep"] = pd.to_datetime(df[col_date], errors="coerce")
    df = df[df["datep"].dt.date >= date_min]
    df = df[df["datep"].dt.date <= date_max]

    out = pd.DataFrame()
    out["datep"] = df["datep"]
    out["supportp"] = df[col_sup].astype(str).str.strip()
    out["support_norm"] = out["supportp"].apply(normalize_support)
    out["heure de diffusion"] = df[col_time]
    out["Marque"] = df[col_mar].astype(str).str.strip()
    out["Marque_norm"] = out["Marque"].apply(brand_key)

    out["Message"] = df[find_column(df_imp, ["message", "storyboard"])] if find_column(df_imp, ["message", "storyboard"]) else None
    out["Produit"] = df[find_column(df_imp, ["produit"])] if find_column(df_imp, ["produit"]) else None
    out["RaisonSociale"] = df[find_column(df_imp, ["raisonsociale", "raison sociale"])] if find_column(df_imp, ["raisonsociale", "raison sociale"]) else None
    out["FormatSec"] = df[find_column(df_imp, ["formatsec", "format"])] if find_column(df_imp, ["formatsec", "format"]) else None
    out["Avant"] = df[find_column(df_imp, ["avant"])] if find_column(df_imp, ["avant"]) else None
    out["Apres"] = df[find_column(df_imp, ["apres", "après"])] if find_column(df_imp, ["apres", "après"]) else None
    out["rangE"] = df[find_column(df_imp, ["range", "rang"])] if find_column(df_imp, ["range", "rang"]) else None
    out["encombE"] = df[find_column(df_imp, ["encombe", "encombrement"])] if find_column(df_imp, ["encombe", "encombrement"]) else None

    out["Code PM"] = None
    out["Commentaire"] = None
    return out

def build_final_df_from_yumi(df_yumi: pd.DataFrame, date_min: date, date_max: date) -> pd.DataFrame:
    col_date   = find_column(df_yumi, ["date"])
    col_chaine = find_column(df_yumi, ["chaine", "chaîne", "support"])
    col_hdeb   = find_column(df_yumi, ["h.debut", "h début", "heure debut", "hdeb", "début"])
    col_marque = find_column(df_yumi, ["marque"])
    if not all([col_date, col_chaine, col_hdeb, col_marque]):
        raise ValueError("DATA YUMI: colonnes minimales manquantes (Date/Chaîne/H.Début/Marque).")

    df = df_yumi.copy()
    df["__Date__"] = pd.to_datetime(df[col_date], errors="coerce")
    df = df[df["__Date__"].dt.date >= date_min]
    df = df[df["__Date__"].dt.date <= date_max]

    out = pd.DataFrame()
    for col in FINAL_COLUMNS_YUMI:
        if col == "Date":
            out[col] = df["__Date__"]
        elif col == "Chaîne":
            out[col] = df[col_chaine].astype(str).str.strip()
        elif col == "H.Début":
            out[col] = df[col_hdeb]
        else:
            if col in df.columns:
                out[col] = df[col]
            else:
                fallback = find_column(df, [col])
                out[col] = df[fallback] if fallback else None

    out["support_norm"] = out["Chaîne"].apply(normalize_support)
    out["Marque_norm"] = out["Marque"].apply(brand_key)
    out["Code Ecran PM"] = None
    out["Commentaire"] = None
    return out

# =========================
# Fill IMPERIUM per client (NO anticipé)
# ✅ AMÉLIORATION DEMANDÉE:
#    - ne pas créer de supports / dates à partir du PM
#    - uniquement supports & dates présents dans la DATA uploadée
# =========================
def fill_codepm_commentaire_per_client(df_client: pd.DataFrame, pm_client: pd.DataFrame, date_min: date, date_max: date):
    df = df_client.copy()
    df["date_only"] = pd.to_datetime(df["datep"], errors="coerce").dt.date
    df["t_real"] = df["heure de diffusion"].apply(to_excel_time)

    col_real_code = find_column(df, ["code ecran", "code écran", "ecran", "écran"])
    if col_real_code:
        df["_real_code_digits"] = df[col_real_code].apply(code_hhmm_digits)
    else:
        df["_real_code_digits"] = None

    if pm_client is None or pm_client.empty:
        base = df.copy()
        for col in FINAL_COLUMNS_IMPERIUM:
            if col not in base.columns:
                base[col] = None
        base.drop(columns=["_real_code_digits"], inplace=True, errors="ignore")
        return base[FINAL_COLUMNS_IMPERIUM]

    pm = pm_client.copy()
    pm = pm[pm["date_only"].notna()]
    pm = pm[(pm["date_only"] >= date_min) & (pm["date_only"] <= date_max)]

    out_all = []
    backlog_by_support = {}  # support_norm -> list[int tv_minutes]

    # ✅ SUPPORTS = UNIQUEMENT ceux présents dans la DATA uploadée
    supports_real = set(df["support_norm"].dropna().unique())
    all_supports = sorted(list(supports_real))

    def insert_non_diffuse_row(dte, sup_disp, codepm, pm_seq=None):
        row = {c: None for c in FINAL_COLUMNS_IMPERIUM}
        row["datep"] = dte
        row["supportp"] = sup_disp
        row["Code PM"] = codepm
        row["Commentaire"] = "Non diffusé"
        if pm_seq is not None:
            row["_pm_seq"] = pm_seq  # ordre PM
        return row

    for sn in all_supports:
        backlog_by_support.setdefault(sn, [])

        real_s = df[df["support_norm"] == sn].copy()
        pm_s = pm[pm["support_norm"] == sn].copy()

        sup_display = str(real_s.iloc[0]["supportp"]) if not real_s.empty else str(sn)

        # ✅ DATES = UNIQUEMENT celles présentes dans la DATA uploadée
        dates_real = set(real_s["date_only"].dropna().unique())
        all_dates = sorted(list(dates_real))

        for d in all_dates:
            if d < date_min or d > date_max:
                continue

            real_day = real_s[real_s["date_only"] == d].copy()
            real_day["_rt"] = real_day["t_real"].apply(lambda t: real_tv_minutes(t))
            real_day = real_day.sort_values("_rt", na_position="last").drop(columns=["_rt"], errors="ignore")

            pm_day = pm_s[pm_s["date_only"] == d].copy()
            if not pm_day.empty:
                pm_day["_PM_TV_MIN_TVDAY"] = pm_day.apply(lambda r: pm_tv_minutes_tvday(r.get("Code PM"), r.get("PM_TV_MIN")), axis=1)
                pm_day = pm_day.sort_values("_PM_TV_MIN_TVDAY", na_position="last").drop(columns=["_PM_TV_MIN_TVDAY"], errors="ignore")

            rt_minutes = [real_tv_minutes(t) if t is not None else None for t in real_day["t_real"].tolist()]
            pm_minutes = [pm_tv_minutes_tvday(pm_day.iloc[j]["Code PM"], pm_day.iloc[j].get("PM_TV_MIN")) for j in range(len(pm_day))]

            real_codes = real_day["_real_code_digits"].tolist() if "_real_code_digits" in real_day.columns else [None] * len(real_day)
            pm_codes = [code_hhmm_digits(v) for v in pm_day["Code PM"].tolist()] if not pm_day.empty else []

            filled_rows = []
            inserted_rows = []

            assign = match_day_exact_then_order_swap(rt_minutes, pm_minutes, real_codes, pm_codes)
            used_pm = set(j for j in assign if j is not None)

            for i in range(len(real_day)):
                r = real_day.iloc[i].copy()
                j = assign[i] if i < len(assign) else None
                if j is not None and j < len(pm_day):
                    pm_row = pm_day.iloc[j]
                    r["Code PM"] = pm_row["Code PM"]
                    diff = None
                    if rt_minutes[i] is not None:
                        pm_min = pm_tv_minutes_tvday(pm_row.get("Code PM"), pm_row.get("PM_TV_MIN"))
                        if pm_min < 10**9:
                            diff = abs(rt_minutes[i] - pm_min)
                    r["Commentaire"] = "Décalage" if (diff is not None and diff > DECALAGE_MINUTES) else None
                else:
                    r["Code PM"] = None
                    # ✅ Compensation seulement si proche d'un PM backlog
                    if backlog_by_support[sn] and rt_minutes[i] is not None:
                        diffs = [abs(rt_minutes[i] - b) for b in backlog_by_support[sn]]
                        best_k = int(pd.Series(diffs).idxmin()) if diffs else None
                        if best_k is not None and diffs[best_k] <= COMPENSATION_MAX_MINUTES:
                            r["Commentaire"] = "Compensation"
                            backlog_by_support[sn].pop(best_k)
                        else:
                            r["Commentaire"] = "Passage supplémentaire"
                    else:
                        r["Commentaire"] = "Passage supplémentaire"
                filled_rows.append(r)

            # remaining pm => Non diffusé (mais uniquement pour dates présentes dans DATA)
            remaining_pm = [idx for idx in range(len(pm_day)) if idx not in used_pm]
            for j in remaining_pm:
                inserted_rows.append(insert_non_diffuse_row(d, sup_display, pm_day.iloc[j]["Code PM"], pm_seq=j))
                pm_min = pm_tv_minutes_tvday(pm_day.iloc[j]["Code PM"], pm_day.iloc[j].get("PM_TV_MIN"))
                if pm_min < 10**9:
                    backlog_by_support[sn].append(pm_min)

            df_filled = pd.DataFrame(filled_rows) if filled_rows else pd.DataFrame()
            df_insert = pd.DataFrame(inserted_rows) if inserted_rows else pd.DataFrame(columns=FINAL_COLUMNS_IMPERIUM)

            def sort_key_tv(row):
                t = to_excel_time(row.get("heure de diffusion"))
                if t is not None:
                    return real_tv_minutes(t)
                return pm_tv_minutes_tvday(row.get("Code PM"), None)

            if not df_filled.empty:
                df_filled["_sort_t"] = df_filled.apply(lambda rr: sort_key_tv(rr), axis=1)
                df_filled["supportp"] = sup_display
                if "_pm_seq" not in df_filled.columns:
                    df_filled["_pm_seq"] = 10**9
            if not df_insert.empty:
                df_insert["_sort_t"] = df_insert.apply(lambda rr: sort_key_tv(rr), axis=1)
                if "_pm_seq" not in df_insert.columns:
                    df_insert["_pm_seq"] = 10**9

            out_day = pd.concat([x for x in [df_filled, df_insert] if not x.empty], ignore_index=True) \
                     if (not df_filled.empty or not df_insert.empty) else pd.DataFrame(columns=FINAL_COLUMNS_IMPERIUM)

            if not out_day.empty:
                out_day = out_day.sort_values(["_sort_t", "_pm_seq"], na_position="last")
                out_day = out_day.drop(columns=["_sort_t", "_pm_seq"], errors="ignore")

            out_all.append(out_day[FINAL_COLUMNS_IMPERIUM])

    out_df = pd.concat(out_all, ignore_index=True) if out_all else df[FINAL_COLUMNS_IMPERIUM].copy()
    out_df.drop(columns=["_real_code_digits"], inplace=True, errors="ignore")
    return out_df[FINAL_COLUMNS_IMPERIUM]

# =========================
# Fill YUMI per client (NO anticipé)
# =========================
def fill_codeecranpm_commentaire_per_client_yumi(df_client: pd.DataFrame, pm_client: pd.DataFrame, date_min: date, date_max: date):
    df = df_client.copy()
    df["date_only"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    df["t_real"] = df["H.Début"].apply(to_excel_time)
    df["_real_code_digits"] = df["Code Ecran"].apply(code_hhmm_digits) if "Code Ecran" in df.columns else None

    if pm_client is None or pm_client.empty:
        base = df.copy()
        for col in FINAL_COLUMNS_YUMI:
            if col not in base.columns:
                base[col] = None
        base.drop(columns=["_real_code_digits"], inplace=True, errors="ignore")
        return base[FINAL_COLUMNS_YUMI]

    pm = pm_client.copy()
    pm = pm[pm["date_only"].notna()]
    pm = pm[(pm["date_only"] >= date_min) & (pm["date_only"] <= date_max)]

    out_all = []
    backlog_by_support = {}

    supports_real = set(df["support_norm"].dropna().unique())
    supports_pm = set(pm["support_norm"].dropna().unique())
    all_supports = sorted(list(supports_real | supports_pm))

    # ✅ YUMI: uniquement 2M & ALAOULA
    all_supports = [sn for sn in all_supports if sn in ALLOWED_YUMI]

    def insert_minimal_row_yumi(dte, chaine, codepm, pm_seq=None):
        row = {c: None for c in FINAL_COLUMNS_YUMI}
        dts = pd.to_datetime(dte)
        row["Date"] = dts
        row["Chaîne"] = chaine
        row["N° Mois"] = int(dts.month) if not pd.isna(dts) else None
        row["Année"] = int(dts.year) if not pd.isna(dts) else None
        row["Code Ecran PM"] = codepm
        row["Commentaire"] = "Non diffusé"
        if pm_seq is not None:
            row["_pm_seq"] = pm_seq
        return row

    for sn in all_supports:
        backlog_by_support.setdefault(sn, [])

        real_s = df[df["support_norm"] == sn].copy()
        pm_s = pm[pm["support_norm"] == sn].copy()

        sup_display = str(real_s.iloc[0]["Chaîne"]) if not real_s.empty else (str(pm_s.iloc[0]["supportp"]) if not pm_s.empty else str(sn))

        dates_real = set(real_s["date_only"].dropna().unique())
        dates_pm = set(pm_s["date_only"].dropna().unique())
        all_dates = sorted(list(dates_real | dates_pm))

        for d in all_dates:
            if d < date_min or d > date_max:
                continue

            real_day = real_s[real_s["date_only"] == d].copy()
            real_day["_rt"] = real_day["t_real"].apply(lambda t: real_tv_minutes(t))
            real_day = real_day.sort_values("_rt", na_position="last").drop(columns=["_rt"], errors="ignore")

            pm_day = pm_s[pm_s["date_only"] == d].copy()
            if not pm_day.empty:
                pm_day["_PM_TV_MIN_TVDAY"] = pm_day.apply(lambda r: pm_tv_minutes_tvday(r.get("Code PM"), r.get("PM_TV_MIN")), axis=1)
                pm_day = pm_day.sort_values("_PM_TV_MIN_TVDAY", na_position="last").drop(columns=["_PM_TV_MIN_TVDAY"], errors="ignore")

            rt_minutes = [real_tv_minutes(t) if t is not None else None for t in real_day["t_real"].tolist()]
            pm_minutes = [pm_tv_minutes_tvday(pm_day.iloc[j]["Code PM"], pm_day.iloc[j].get("PM_TV_MIN")) for j in range(len(pm_day))]

            real_codes = real_day["_real_code_digits"].tolist() if "_real_code_digits" in real_day.columns else [None] * len(real_day)
            pm_codes = [code_hhmm_digits(v) for v in pm_day["Code PM"].tolist()] if not pm_day.empty else []

            filled_rows = []
            inserted_rows = []

            if len(real_day) == 0 and len(pm_day) > 0:
                for j, (_, p) in enumerate(pm_day.iterrows()):
                    inserted_rows.append(insert_minimal_row_yumi(d, sup_display, p["Code PM"], pm_seq=j))
                    pm_min = pm_tv_minutes_tvday(p["Code PM"], p.get("PM_TV_MIN"))
                    if pm_min < 10**9:
                        backlog_by_support[sn].append(pm_min)
            else:
                assign = match_day_exact_then_order_swap(rt_minutes, pm_minutes, real_codes, pm_codes)
                used_pm = set(j for j in assign if j is not None)

                for i in range(len(real_day)):
                    r = real_day.iloc[i].copy()
                    j = assign[i] if i < len(assign) else None
                    if j is not None and j < len(pm_day):
                        pm_row = pm_day.iloc[j]
                        r["Code Ecran PM"] = pm_row["Code PM"]

                        diff = None
                        if rt_minutes[i] is not None:
                            pm_min = pm_tv_minutes_tvday(pm_row.get("Code PM"), pm_row.get("PM_TV_MIN"))
                            if pm_min < 10**9:
                                diff = abs(rt_minutes[i] - pm_min)

                        r["Commentaire"] = "Décalage" if (diff is not None and diff > DECALAGE_MINUTES) else None
                    else:
                        r["Code Ecran PM"] = None
                        if backlog_by_support[sn] and rt_minutes[i] is not None:
                            diffs = [abs(rt_minutes[i] - b) for b in backlog_by_support[sn]]
                            best_k = int(pd.Series(diffs).idxmin()) if diffs else None
                            if best_k is not None and diffs[best_k] <= COMPENSATION_MAX_MINUTES:
                                r["Commentaire"] = "Compensation"
                                backlog_by_support[sn].pop(best_k)
                            else:
                                r["Commentaire"] = "Passage supplémentaire"
                        else:
                            r["Commentaire"] = "Passage supplémentaire"
                    filled_rows.append(r)

                remaining_pm = [idx for idx in range(len(pm_day)) if idx not in used_pm]
                for j in remaining_pm:
                    inserted_rows.append(insert_minimal_row_yumi(d, sup_display, pm_day.iloc[j]["Code PM"], pm_seq=j))
                    pm_min = pm_tv_minutes_tvday(pm_day.iloc[j]["Code PM"], pm_day.iloc[j].get("PM_TV_MIN"))
                    if pm_min < 10**9:
                        backlog_by_support[sn].append(pm_min)

            df_filled = pd.DataFrame(filled_rows) if filled_rows else pd.DataFrame()
            df_insert = pd.DataFrame(inserted_rows) if inserted_rows else pd.DataFrame(columns=FINAL_COLUMNS_YUMI)

            def sort_key_tv_yumi(row):
                t = to_excel_time(row.get("H.Début"))
                if t is not None:
                    return real_tv_minutes(t)
                return pm_tv_minutes_tvday(row.get("Code Ecran PM"), None)

            if not df_filled.empty:
                df_filled["_sort_t"] = df_filled.apply(lambda rr: sort_key_tv_yumi(rr), axis=1)
                df_filled["Chaîne"] = sup_display
                if "_pm_seq" not in df_filled.columns:
                    df_filled["_pm_seq"] = 10**9
            if not df_insert.empty:
                df_insert["_sort_t"] = df_insert.apply(lambda rr: sort_key_tv_yumi(rr), axis=1)
                if "_pm_seq" not in df_insert.columns:
                    df_insert["_pm_seq"] = 10**9

            out_day = pd.concat([x for x in [df_filled, df_insert] if not x.empty], ignore_index=True) \
                     if (not df_filled.empty or not df_insert.empty) else pd.DataFrame(columns=FINAL_COLUMNS_YUMI)

            if not out_day.empty:
                out_day = out_day.sort_values(["_sort_t", "_pm_seq"], na_position="last")
                out_day = out_day.drop(columns=["_sort_t", "_pm_seq"], errors="ignore")

            out_all.append(out_day[FINAL_COLUMNS_YUMI])

    out_df = pd.concat(out_all, ignore_index=True) if out_all else df.copy()
    out_df.drop(columns=["_real_code_digits"], inplace=True, errors="ignore")
    return out_df[FINAL_COLUMNS_YUMI]

# =========================
# Styles + finalize
# =========================
def apply_row_style_from_template(style_row_cells, ws, row_idx, final_cols):
    for c in range(1, len(final_cols) + 1):
        src = style_row_cells[c - 1]
        dst = ws.cell(row_idx, c)
        dst._style = pycopy(src._style)
        dst.number_format = src.number_format
        dst.font = pycopy(src.font)
        dst.fill = pycopy(src.fill)
        dst.border = pycopy(src.border)
        dst.alignment = pycopy(src.alignment)
        dst.protection = pycopy(src.protection)

def finalize_sheet(ws, style_row_cells, final_cols, total_col_name: str, mode: str):
    ws["A6"].value = None
    ws["B4"].value = date.today()
    ws["B4"].number_format = "dd/mm/yyyy"

    # ✅ B5: YUMI = F10 ; Imperium = H10
    if mode == "Suivi YUMI":
        ws["B5"].value = ws["F10"].value
    else:
        ws["B5"].value = ws["H10"].value

    total_col_idx = final_cols.index(total_col_name) + 1 if total_col_name in final_cols else 4

    last_data_row = 9
    for r in range(ws.max_row, DATA_START_ROW - 1, -1):
        if (ws.cell(r, 1).value not in (None, "")) or (ws.cell(r, 2).value not in (None, "")) or (ws.cell(r, total_col_idx).value not in (None, "")):
            last_data_row = max(r, DATA_START_ROW)
            break
    if last_data_row < DATA_START_ROW:
        last_data_row = DATA_START_ROW

    count_vals = 0
    for r in range(DATA_START_ROW, last_data_row + 1):
        if ws.cell(r, total_col_idx).value not in (None, ""):
            count_vals += 1

    total_row = last_data_row + 1
    ws.insert_rows(total_row)
    apply_row_style_from_template(style_row_cells, ws, total_row, final_cols)

    ws.cell(total_row, 1).value = "Total"
    ws.cell(total_row, 2).value = count_vals

    grey_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in (1, 2):
        cell = ws.cell(total_row, col)
        cell.fill = grey_fill
        cell.border = thin_border
        cell.font = pycopy(cell.font)
        cell.font = cell.font.copy(bold=True)

    empty_border = Border()
    empty_fill = PatternFill(fill_type=None)
    for col in range(3, len(final_cols) + 1):
        cell = ws.cell(total_row, col)
        cell.value = None
        cell.border = empty_border
        cell.fill = empty_fill

# =========================
# Build workbooks
# =========================
def build_client_workbook_from_template(template_wb: openpyxl.Workbook, client_name: str, df_client: pd.DataFrame, final_cols: list[str], mode: str) -> bytes:
    wb = template_wb
    template_ws = wb.worksheets[0]
    style_row_cells = [template_ws.cell(DATA_START_ROW, c) for c in range(1, len(final_cols) + 1)]

    def reset_sheet(ws):
        if ws.max_row > DATA_START_ROW:
            ws.delete_rows(DATA_START_ROW + 1, ws.max_row - DATA_START_ROW)
        for c in range(1, len(final_cols) + 1):
            ws.cell(DATA_START_ROW, c).value = None
        for c, col in enumerate(final_cols, start=1):
            ws.cell(HEADER_ROW, c).value = col
        ws.sheet_view.showGridLines = False

    reset_sheet(template_ws)

    if mode == "Suivi Imperium":
        supports = list(df_client["supportp"].dropna().unique())
        get_sub = lambda sup: df_client[df_client["supportp"] == sup].copy()
        sheet_title = lambda sup: safe_sheet_name(f"{client_name} - {str(sup).strip()}")
        total_col_name = "Code PM"
    else:
        supports = [s for s in df_client["Chaîne"].dropna().unique()
                    if normalize_support(s) in ALLOWED_YUMI]
        get_sub = lambda sup: df_client[df_client["Chaîne"] == sup].copy()
        sheet_title = lambda sup: safe_sheet_name(f"{client_name} - {str(sup).strip()}")
        total_col_name = "Code Ecran PM"

    if not supports:
        supports = ["Support"]

    for sup in supports:
        ws = wb.copy_worksheet(template_ws)
        ws.title = sheet_title(sup)
        reset_sheet(ws)

        sub = get_sub(sup)

        for i in range(len(sub)):
            r_idx = DATA_START_ROW + i
            if r_idx > DATA_START_ROW:
                ws.insert_rows(r_idx)
                apply_row_style_from_template(style_row_cells, ws, r_idx, final_cols)

            for c, col in enumerate(final_cols, start=1):
                val = sub.iloc[i][col] if col in sub.columns else None
                if col in ("heure de diffusion", "H.Début", "H.Fin"):
                    val = to_excel_time(val)
                ws.cell(r_idx, c).value = val

        finalize_sheet(ws, style_row_cells, final_cols, total_col_name=total_col_name, mode=mode)

    wb.remove(template_ws)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# =========================
# UI
# =========================
st.title("📊 Suivi Pige — Automatisation (PM 2026 unique)")
mode = st.radio("Type de suivi", ["Suivi Imperium", "Suivi YUMI"], horizontal=True)
st.caption("PM unique : 1 feuille = 1 client + 1 chaîne | Col A=Date | Col B=Ecran")

template_ok = False
try:
    _ = load_template_workbook(mode)
    template_ok = True
    st.success("Template OK ✅")
except Exception as e:
    st.error(f"Template introuvable ❌ : {e}")

data_in = st.file_uploader("1) Uploader DATA IMPERIUM" if mode == "Suivi Imperium" else "1) Uploader DATA YUMI", type=["xlsx"])
pm_file = st.file_uploader("2) Uploader PM 2026 (1 fichier)", type=["xlsx"])
max_date_ui = st.date_input("3) Date max (N-1 par défaut)", value=date.today() - timedelta(days=1))

if st.button("Lancer la génération", use_container_width=True, disabled=(not template_ok)):
    if not data_in:
        st.warning("Upload DATA.")
    elif not pm_file:
        st.warning("Upload PM 2026.xlsx.")
    else:
        try:
            with st.spinner("Génération en cours..."):
                df_in = pd.read_excel(data_in)

                min_raw, max_raw = get_min_max_date_from_raw(df_in, mode)
                if min_raw is None or max_raw is None:
                    raise ValueError("Impossible de détecter min/max date dans la data brute.")

                date_min = min_raw
                date_max = min(max_date_ui, max_raw)
                st.info(f"Fenêtre dates utilisée: {date_min} → {date_max} (bornée par la data brute)")

                if mode == "Suivi Imperium":
                    df_all = build_final_df_from_imperium(df_in, date_min=date_min, date_max=date_max)
                    final_cols = FINAL_COLUMNS_IMPERIUM
                    client_col = "Marque"
                else:
                    df_all = build_final_df_from_yumi(df_in, date_min=date_min, date_max=date_max)
                    final_cols = FINAL_COLUMNS_YUMI
                    client_col = "Marque"
                    df_all = df_all[df_all["support_norm"].isin(ALLOWED_YUMI)].copy()

                known_supports = set(df_all["support_norm"].dropna().unique()) | {"2M", "MBC5", "ALAOULA"}
                pmv_all = read_pm_2026_workbook(pm_file.getvalue(), known_supports)

                client_files = {}
                for client_name in sorted(df_all[client_col].dropna().unique()):
                    df_client_raw = df_all[df_all[client_col] == client_name].copy()
                    client_norm = brand_key(client_name)

                    pm_client = pmv_all[pmv_all["PM_FILE_BRAND_N"] == client_norm].copy()
                    if pm_client.empty and not pmv_all.empty:
                        pm_client = pmv_all[
                            pmv_all["PM_FILE_BRAND_N"].apply(lambda x: (x in client_norm) or (client_norm in x))
                        ].copy()

                    if mode == "Suivi Imperium":
                        df_client_done = fill_codepm_commentaire_per_client(df_client_raw, pm_client, date_min=date_min, date_max=date_max)
                    else:
                        df_client_done = fill_codeecranpm_commentaire_per_client_yumi(df_client_raw, pm_client, date_min=date_min, date_max=date_max)

                    df_client_done = df_client_done.copy()
                    for helper in ("support_norm", "Marque_norm", "date_only", "t_real", "_real_code_digits", "_pm_seq"):
                        df_client_done.drop(columns=[helper], inplace=True, errors="ignore")

                    template_wb = load_template_workbook(mode)
                    xlsx_bytes = build_client_workbook_from_template(template_wb, client_name, df_client_done, final_cols, mode=mode)
                    client_files[f"Suivi_{client_name}.xlsx"] = xlsx_bytes

                st.session_state.client_files = client_files
                st.session_state.zip_bytes = make_zip(client_files)
                st.session_state.last_run_info = f"{len(client_files)} fichiers générés"

        except Exception as e:
            st.error(f"Erreur: {e}")

if st.session_state.client_files:
    st.success(st.session_state.last_run_info)
    st.download_button(
        "📦 Télécharger ZIP",
        data=st.session_state.zip_bytes,
        file_name=f"Suivis_{mode.replace(' ', '_')}_{max_date_ui.isoformat()}.zip",
        mime="application/zip",
        use_container_width=True
    )
    st.divider()
    cols = st.columns(3)
    for i, (fname, data) in enumerate(st.session_state.client_files.items()):
        with cols[i % 3]:
            st.download_button(
                f"📥 {fname}",
                data=data,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
