"""
PPT Media Review Engine v10
============================
Architecture DYNAMIQUE :
- Slides 1, 2, 3 : toujours présentes (titre, overview, annonceurs)
- Slides médias : générées uniquement pour les médias présents dans la data
- Chaque slide média est clonée depuis le template puis adaptée
- Les slides absentes sont complètement supprimées du PPT final

Médias supportés : AF, TV, RD, PR, CN
Ordre de sortie  : AF → TV → RD → PR → CN (si présents)
"""

import io
import zipfile
import re
import math
from copy import deepcopy
from typing import Optional

import pandas as pd
from lxml import etree

# ─────────────────────────────────────────────
# NAMESPACES
# ─────────────────────────────────────────────
CNS  = "http://schemas.openxmlformats.org/drawingml/2006/chart"
ANS  = "http://schemas.openxmlformats.org/drawingml/2006/main"
PNS  = "http://schemas.openxmlformats.org/presentationml/2006/main"
RNS  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CTNS = "http://schemas.openxmlformats.org/package/2006/content-types"

def ctag(n): return f"{{{CNS}}}{n}"
def atag(n): return f"{{{ANS}}}{n}"
def ptag(n): return f"{{{PNS}}}{n}"

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
MEDIA_LABELS = {
    "AF": "Affichage (OOH)",
    "TV": "TV",
    "PR": "Presse",
    "RD": "Radio",
    "CN": "Cinéma",
}
MEDIA_SHORT = {"AF": "OOH", "TV": "TV", "PR": "PR", "RD": "RD", "CN": "CN"}
MEDIA_SUP_LABEL = {"AF": "ville", "TV": "chaîne", "PR": "support", "RD": "station", "CN": "support"}

# Ordre d'affichage des slides médias dans le PPT final
MEDIA_ORDER = ["AF", "TV", "RD", "PR", "CN"]

# Mapping slide template → média + charts (rIds dans slide rels)
# slide_num: (media_code, chart_annual_rid, chart_pie_rid, chart_topann_rid)
TEMPLATE_SLIDE_MAP = {
    4: ("AF", "rId2", "rId4", "rId3"),   # chart7=annual(rId2), chart9=pie(rId4), chart8=topAnn(rId3)
    5: ("RD", "rId4", "rId2", "rId3"),   # chart12=annual(rId4), chart10=pie(rId2), chart11=topAnn(rId3)
    6: ("TV", "rId3", "rId2", "rId4"),   # chart14=annual(rId3), chart13=pie(rId2), chart15=topAnn(rId4)
    7: ("CN", "rId2", "rId3", "rId4"),   # chart16=annual(rId2), chart17=pie(rId3), chart18=topAnn(rId4)
    8: ("PR", "rId4", "rId2", "rId3"),   # chart21=annual(rId4), chart19=pie(rId2), chart20=topAnn(rId3)
}

# Shape textuelle à injecter par slide média
# "comment_shape" = shape qui reçoit le commentaire IA (varie selon slide)
MEDIA_SLIDE_SHAPES = {
    4: {"comment": "TextBox 15"},   # AF
    5: {"comment": "TextBox 15"},   # RD
    6: {"comment": "Rectangle 3"},  # TV  — shape différente !
    7: {"comment": "TextBox 15"},   # CN
    8: {"comment": "TextBox 15"},   # PR
}

MONTHS_FR = ["Jan","Fév","Mar","Avr","Mai","Juin","Juil","Août","Sep","Oct","Nov","Déc"]
TOP_N_ANNONCEURS = 15
TOP_N_SUPPORT    = 12
FMT_MILLIONS = "0.0"
FMT_PERCENT  = "0%"


# ═══════════════════════════════════════════════════════════════
# 1. MOTEUR DE CALCUL
# ═══════════════════════════════════════════════════════════════
class MediaCalculator:
    def __init__(self, df: pd.DataFrame, secteur_filter: str, sous_secteur_filter: Optional[str]):
        self.raw = df.copy()
        self.raw.columns = [c.strip() for c in self.raw.columns]
        tarif_col = next((c for c in self.raw.columns if "tarif" in c.lower()), None)
        if not tarif_col:
            raise ValueError("Colonne tarif introuvable.")
        self.raw.rename(columns={tarif_col: "tarif"}, inplace=True)
        self.raw["tarif"] = pd.to_numeric(self.raw["tarif"], errors="coerce").fillna(0)

        df_f = self.raw.copy()
        if secteur_filter:
            df_f = df_f[df_f["Secteur"].str.strip() == secteur_filter.strip()]
        if sous_secteur_filter:
            df_f = df_f[df_f["SousSecteur"].str.strip() == sous_secteur_filter.strip()]
        self.df = df_f.copy()

        self.years      = sorted(self.df["Anp"].dropna().unique().tolist())
        self.year_last  = self.years[-1] if self.years else None
        self.year_prev  = self.years[-2] if len(self.years) >= 2 else None
        self.year_prev2 = self.years[-3] if len(self.years) >= 3 else None

        # Médias présents : détectés directement depuis la data uploadée (investissement > 0)
        # Les 5 médias possibles sont : AF, TV, RD, PR, CN
        # Ordre de sortie dans le PPT : AF → TV → RD → PR → CN
        medias_in_data = set(
            self.df[self.df["tarif"] > 0]["media"].dropna().unique().tolist()
        )
        self.medias_present = [m for m in MEDIA_ORDER if m in medias_in_data]

    def total_by_year(self) -> dict:
        return self.df.groupby("Anp")["tarif"].sum().reindex(self.years, fill_value=0).to_dict()

    def total_by_year_media(self) -> pd.DataFrame:
        pt = self.df.groupby(["Anp", "media"])["tarif"].sum().unstack(fill_value=0)\
               .reindex(self.years, fill_value=0)
        for m in MEDIA_ORDER:
            if m not in pt.columns:
                pt[m] = 0
        return pt

    def media_mix_last_year(self) -> dict:
        if not self.year_last:
            return {}
        row = self.total_by_year_media().loc[self.year_last]
        total = row.sum()
        return {k: v / total * 100 for k, v in row.items() if v > 0} if total else {}

    def seasonality_by_year(self) -> pd.DataFrame:
        pt = self.df.groupby(["Anp", "moisp"])["tarif"].sum()\
               .unstack(fill_value=0).reindex(columns=range(1, 13), fill_value=0)
        pt.index = [int(y) for y in pt.index]
        return pt

    def top_annonceurs_by_year(self, year, n=TOP_N_ANNONCEURS) -> pd.Series:
        return self.df[self.df["Anp"] == year].groupby("Marque")["tarif"]\
               .sum().sort_values(ascending=False).head(n)

    def _focus(self, code):
        return self.df[self.df["media"] == code].copy()

    def total_focus_by_year(self, code) -> dict:
        return self._focus(code).groupby("Anp")["tarif"].sum()\
               .reindex(self.years, fill_value=0).to_dict()

    def top_ann_focus_last(self, code, n=TOP_N_ANNONCEURS) -> pd.Series:
        f = self._focus(code)
        return f[f["Anp"] == self.year_last].groupby("Marque")["tarif"]\
               .sum().sort_values(ascending=False).head(n)

    def split_support_last(self, code, n=TOP_N_SUPPORT, seuil_pct=3.0) -> pd.Series:
        f = self._focus(code)
        s = f[f["Anp"] == self.year_last].groupby("supportp")["tarif"]\
              .sum().sort_values(ascending=False)
        if s.empty:
            return s
        total = s.sum()
        mask_small = (s / total * 100) < seuil_pct
        main = s[~mask_small].head(n)
        autres_val = s[mask_small].sum()
        if autres_val > 0:
            main = pd.concat([main, pd.Series({"Autres": autres_val})])
        return main

    def evol_pct(self, new, old):
        return (new - old) / abs(old) * 100 if old and old != 0 else None

    def sos(self, code=None) -> dict:
        top   = self.top_ann_focus_last(code) if code else self.top_annonceurs_by_year(self.year_last)
        total = self.total_focus_by_year(code).get(self.year_last, 0) if code \
                else self.total_by_year().get(self.year_last, 0)
        return {b: v / total * 100 for b, v in top.items()} if total else {}

    def summary_stats(self) -> dict:
        totals = self.total_by_year()
        mix    = self.media_mix_last_year()
        s = {
            "years": self.years, "year_last": self.year_last, "totals": totals,
            "total_last":  totals.get(self.year_last, 0),
            "total_prev":  totals.get(self.year_prev, 0),
            "total_prev2": totals.get(self.year_prev2, 0),
            "evol_vs_prev":  self.evol_pct(totals.get(self.year_last, 0), totals.get(self.year_prev, 0)),
            "evol_vs_prev2": self.evol_pct(totals.get(self.year_last, 0), totals.get(self.year_prev2, 0)),
            "media_mix": mix,
            "dominant_media":     max(mix, key=mix.get) if mix else None,
            "dominant_media_pct": max(mix.values()) if mix else 0,
        }
        seas = self.seasonality_by_year()
        if self.year_last and self.year_last in seas.index:
            peak = int(seas.loc[self.year_last].idxmax())
            s["peak_month"] = MONTHS_FR[peak - 1]
            s["peak_value"] = seas.loc[self.year_last, peak]

        top_ann = self.top_annonceurs_by_year(self.year_last) if self.year_last else pd.Series()
        if len(top_ann):
            s["top1_ann"] = top_ann.index[0]
            s["top1_val"] = top_ann.iloc[0]
            s["top1_sos"] = self.sos().get(top_ann.index[0], 0)
        if len(top_ann) > 1:
            s["top3_sos"] = sum(list(self.sos().values())[:3])

        for code in MEDIA_ORDER:
            sub = self.total_focus_by_year(code)
            if sub.get(self.year_last, 0) > 0:
                s[f"{code}_last"] = sub.get(self.year_last, 0)
                s[f"{code}_prev"] = sub.get(self.year_prev, 0)
                s[f"{code}_evol"] = self.evol_pct(sub.get(self.year_last, 0), sub.get(self.year_prev, 0))
                sup = self.split_support_last(code)
                if len(sup):
                    tot = sub.get(self.year_last, 0)
                    s[f"{code}_top_sup"]     = sup.index[0].replace("_", " ").upper()
                    s[f"{code}_top_sup_pct"] = sup.iloc[0] / tot * 100 if tot else 0
                ann = self.top_ann_focus_last(code)
                if len(ann):
                    s[f"{code}_top1_ann"] = ann.index[0]
                    s[f"{code}_top1_sos"] = self.sos(code).get(ann.index[0], 0)
        return s


# ═══════════════════════════════════════════════════════════════
# 2. COMMENTAIRES IA
# ═══════════════════════════════════════════════════════════════
def generate_comments_via_claude(stats: dict, secteur: str, label: str, api_key: str) -> dict:
    import requests, json, re as re2

    yl  = stats.get("year_last", "")
    yp  = stats["years"][-2] if len(stats["years"]) >= 2 else ""
    yp2 = stats["years"][-3] if len(stats["years"]) >= 3 else ""

    def fm(v):  return f"{v/1e6:.1f} M MAD" if v else "N/A"
    def fp(v):  return f"{'+'if v and v>0 else ''}{v:.1f}%" if v is not None else "N/A"
    def ypl(y): return str(int(y)) if y else "N-1"

    medias_present = [c for c in MEDIA_ORDER if stats.get(f"{c}_last")]

    ctx = (
        f"Secteur: {secteur} | Sous-secteur: {label} | Période: {yp2}–{yl}\n"
        f"GLOBAL: {yl}: {fm(stats.get('total_last'))} | {ypl(yp)}: {fm(stats.get('total_prev'))} | {ypl(yp2)}: {fm(stats.get('total_prev2'))}\n"
        f"Évol vs {ypl(yp)}: {fp(stats.get('evol_vs_prev'))} | Évol vs {ypl(yp2)}: {fp(stats.get('evol_vs_prev2'))}\n"
        f"Mix: {', '.join(f'{k}:{v:.0f}%' for k,v in stats.get('media_mix',{}).items())} | Pic: {stats.get('peak_month','')} ({fm(stats.get('peak_value'))})\n"
        f"Top annonceur: {stats.get('top1_ann','')} — {fm(stats.get('top1_val'))} — SOS {stats.get('top1_sos',0):.0f}%\n"
    ) + "\n".join(
        f"{c}: {fm(stats.get(f'{c}_last'))} | Évol {fp(stats.get(f'{c}_evol'))} | "
        f"Top: {stats.get(f'{c}_top_sup','')} ({stats.get(f'{c}_top_sup_pct',0):.0f}%) | "
        f"Leader: {stats.get(f'{c}_top1_ann','')} (SOS {stats.get(f'{c}_top1_sos',0):.0f}%)"
        for c in medias_present
    )

    slides_json = '{\n  "slide2_global": "...",\n  "slide2_headline": "...",\n  "slide3_annonceurs": "..."'
    for c in medias_present:
        slides_json += f',\n  "slide_{c.lower()}": "commentaire {MEDIA_LABELS.get(c,c)}"'
    slides_json += "\n}"

    prompt = (
        f"Tu es un expert media planner senior au Maroc avec 15 ans d'expérience.\n"
        f"Génère des commentaires analytiques DÉTAILLÉS pour un Media Review PPT professionnel.\n\n"
        f"Données:\n{ctx}\n\n"
        f"RÈGLES:\n"
        f"- 4 à 6 phrases par commentaire, chiffres exacts en M MAD et %\n"
        f"- Comparer les 3 années, identifier leaders/tendances\n"
        f"- Ton analytique et professionnel\n"
        f"- JAMAIS de formules génériques\n\n"
        f"Retourne UNIQUEMENT un JSON valide (sans markdown, sans backticks):\n{slides_json}"
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": "claude-sonnet-4-6", "max_tokens": 2000,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=30,
        )
        resp.raise_for_status()
        text = re2.sub(r"^```json\s*|```$", "", resp.json()["content"][0]["text"].strip(),
                       flags=re2.MULTILINE).strip()
        return json.loads(text)
    except Exception:
        # Fallback sans IA
        yp_l = ypl(yp); yp2_l = ypl(yp2)
        def fm2(v): return f"{v/1e6:.1f} M MAD" if v else "N/A"
        out = {
            "slide2_global":     f"Mix médias {yl}: {', '.join(f'{MEDIA_LABELS.get(k,k)} {v:.0f}%' for k,v in stats.get('media_mix',{}).items())}. Pic saisonnalité: {stats.get('peak_month','')} ({fm2(stats.get('peak_value'))}).",
            "slide2_headline":   f"{yl}: {fm2(stats.get('total_last'))} ({fp(stats.get('evol_vs_prev'))} vs {yp_l}), ({fp(stats.get('evol_vs_prev2'))} vs {yp2_l})",
            "slide3_annonceurs": f"Leader {yl}: {stats.get('top1_ann','')} avec {fm2(stats.get('top1_val'))} (SOS {stats.get('top1_sos',0):.0f}%). Top 3 = {stats.get('top3_sos',0):.0f}% du marché.",
        }
        for c in medias_present:
            evol = stats.get(f"{c}_evol")
            trend = "en hausse" if evol and evol > 0 else "en baisse"
            top_sup = stats.get(f"{c}_top_sup", "")
            top_pct = stats.get(f"{c}_top_sup_pct", 0)
            top1    = stats.get(f"{c}_top1_ann", "")
            top1_sos = stats.get(f"{c}_top1_sos", 0)
            media_name = MEDIA_LABELS.get(c, c)
            dominance = "position dominante" if top1_sos > 50 else "marché fragmenté"
            trend_txt = ("Tendance positive — regain d'intérêt." if evol and evol > 10
                         else "Tendance négative — réallocations vers autres médias." if evol and evol < -10
                         else "Évolution modérée — marché en stabilisation.")
            out[f"slide_{c.lower()}"] = (
                f"{media_name} {yl}: {fm2(stats.get(f'{c}_last'))} ({fp(evol)} vs {yp_l}). "
                f"Le marché {media_name.lower()} est {trend} sur {yp2_l}–{yl}. "
                f"Support dominant: {top_sup} = {top_pct:.0f}% des invest. {yl}. "
                f"Leader: {top1} SOS {top1_sos:.0f}% ({dominance}). {trend_txt}"
            )
        return out


# ═══════════════════════════════════════════════════════════════
# 3. HELPERS XML — CHARTS
# ═══════════════════════════════════════════════════════════════
def smart_max(values_mad: list) -> float:
    vals = [v for v in values_mad if v is not None and v > 0]
    if not vals:
        return 1.0
    raw_max = max(vals) / 1e6
    magnitude = 10 ** math.floor(math.log10(raw_max))
    for mult in [1, 1.2, 1.5, 2, 2.5, 3, 4, 5, 6, 7, 8, 10, 12, 15, 20, 25, 30, 40, 50]:
        candidate = magnitude * mult
        if candidate >= raw_max * 1.10:
            return candidate
    return raw_max * 1.2


def _fix_val_axis(root, max_val_millions: float):
    for val_ax in root.findall(f".//{ctag('valAx')}"):
        for child in list(val_ax):
            if child.tag == ctag("dispUnits"):
                val_ax.remove(child)
        nf = val_ax.find(ctag("numFmt"))
        if nf is None:
            nf = etree.SubElement(val_ax, ctag("numFmt"))
        nf.set("formatCode", "0.0")
        nf.set("sourceLinked", "0")
        scaling = val_ax.find(ctag("scaling"))
        if scaling is None:
            scaling = etree.Element(ctag("scaling"))
            val_ax.insert(0, scaling)
        max_el = scaling.find(ctag("max"))
        if max_el is None:
            max_el = etree.SubElement(scaling, ctag("max"))
        max_el.set("val", str(round(max_val_millions, 2)))
        min_el = scaling.find(ctag("min"))
        if min_el is None:
            min_el = etree.SubElement(scaling, ctag("min"))
        min_el.set("val", "0")


def _set_num_fmt(dlbls_el, fmt_code: str):
    nf = dlbls_el.find(ctag("numFmt"))
    if nf is None:
        nf = etree.SubElement(dlbls_el, ctag("numFmt"))
        dlbls_el.insert(0, nf)
    nf.set("formatCode", fmt_code)
    nf.set("sourceLinked", "0")


def _set_show_flags(dlbls_el, show_val="0", show_pct="0"):
    for tag, val in [("showVal", show_val), ("showPercent", show_pct),
                     ("showLegendKey", "0"), ("showCatName", "0"),
                     ("showSerName", "0"), ("showBubbleSize", "0")]:
        el = dlbls_el.find(ctag(tag))
        if el is not None:
            el.set("val", val)


def _rebuild_cache(ser_el, categories: list, values: list, divide_by: float = 1.0):
    cat_el = ser_el.find(ctag("cat"))
    if cat_el is not None:
        for ref_type in ["numRef", "strRef"]:
            ref = cat_el.find(ctag(ref_type))
            if ref is not None:
                for cache_type in ["numCache", "strCache"]:
                    cache = ref.find(ctag(cache_type))
                    if cache is not None:
                        for pt in cache.findall(ctag("pt")): cache.remove(pt)
                        pc = cache.find(ctag("ptCount"))
                        if pc is None: pc = etree.SubElement(cache, ctag("ptCount"))
                        pc.set("val", str(len(categories)))
                        for i, c in enumerate(categories):
                            pt = etree.SubElement(cache, ctag("pt"))
                            pt.set("idx", str(i))
                            etree.SubElement(pt, ctag("v")).text = str(c)
                        f_el = ref.find(ctag("f"))
                        if f_el is not None:
                            m = re.match(r"(.+)!\$([A-Z]+)\$\d+:\$[A-Z]+\$\d+", f_el.text or "")
                            if m: f_el.text = f"{m.group(1)}!${m.group(2)}$2:${m.group(2)}${len(categories)+1}"

    val_el = ser_el.find(ctag("val"))
    if val_el is not None:
        num_ref = val_el.find(ctag("numRef"))
        if num_ref is not None:
            cache = num_ref.find(ctag("numCache"))
            if cache is None: cache = etree.SubElement(num_ref, ctag("numCache"))
            for pt in cache.findall(ctag("pt")): cache.remove(pt)
            fc = cache.find(ctag("formatCode"))
            if fc is None: fc = etree.SubElement(cache, ctag("formatCode"))
            fc.text = "0.0"
            pc = cache.find(ctag("ptCount"))
            if pc is None: pc = etree.SubElement(cache, ctag("ptCount"))
            pc.set("val", str(len(values)))
            for i, v in enumerate(values):
                if v is None: continue
                pt = etree.SubElement(cache, ctag("pt"))
                pt.set("idx", str(i))
                v_out = v / divide_by if divide_by != 1.0 else v
                etree.SubElement(pt, ctag("v")).text = str(v_out)
            f_el = num_ref.find(ctag("f"))
            if f_el is not None:
                m = re.match(r"(.+)!\$([A-Z]+)\$\d+:\$[A-Z]+\$\d+", f_el.text or "")
                if m: f_el.text = f"{m.group(1)}!${m.group(2)}$2:${m.group(2)}${len(values)+1}"


def _set_series_name(ser_el, name: str):
    tx_v = ser_el.find(f".//{ctag('tx')}//{ctag('v')}")
    if tx_v is not None:
        tx_v.text = str(name)


def _set_series_color(ser_el, hex_color: str):
    dns = ANS
    spPr = ser_el.find(ctag("spPr"))
    if spPr is None:
        spPr = etree.SubElement(ser_el, ctag("spPr"))
    for fill_tag in [f"{{{dns}}}solidFill", f"{{{dns}}}noFill", f"{{{dns}}}gradFill"]:
        old = spPr.find(fill_tag)
        if old is not None: spPr.remove(old)
    solid = etree.SubElement(spPr, f"{{{dns}}}solidFill")
    srgb = etree.SubElement(solid, f"{{{dns}}}srgbClr")
    srgb.set("val", hex_color)


def _build_peak_dlbls(ser_el, values: list, top_n: int = 3):
    dlbls = ser_el.find(ctag("dLbls"))
    if dlbls is None:
        return
    for dl in dlbls.findall(ctag("dLbl")):
        dlbls.remove(dl)
    _set_num_fmt(dlbls, FMT_MILLIONS)
    sv = dlbls.find(ctag("showVal"))
    if sv is not None: sv.set("val", "0")
    if not values:
        return
    vals_clean = [(i, v) for i, v in enumerate(values) if v is not None and v > 0]
    vals_clean.sort(key=lambda x: -x[1])
    peak_indices = {i for i, _ in vals_clean[:top_n]}
    for idx in sorted(peak_indices):
        dl = etree.Element(ctag("dLbl"))
        etree.SubElement(dl, ctag("idx")).set("val", str(idx))
        nf = etree.SubElement(dl, ctag("numFmt"))
        nf.set("formatCode", FMT_MILLIONS); nf.set("sourceLinked", "0")
        etree.SubElement(dl, ctag("dLblPos")).set("val", "t")
        for tag in ["showLegendKey","showVal","showCatName","showSerName","showPercent","showBubbleSize"]:
            e = etree.SubElement(dl, ctag(tag))
            e.set("val", "1" if tag == "showVal" else "0")
        dlbls.insert(0, dl)


# ─── Fonctions de traitement des charts ──────────────────────
def process_chart_annual(chart_xml: bytes, cats: list, vals: list, label: str) -> bytes:
    root = etree.fromstring(chart_xml)
    sers = root.findall(f".//{ctag('ser')}")
    if sers:
        _rebuild_cache(sers[0], cats, vals, divide_by=1e6)
        _set_series_name(sers[0], label)
        dlbls = root.find(f".//{ctag('dLbls')}")
        if dlbls is not None:
            _set_num_fmt(dlbls, FMT_MILLIONS)
            _set_show_flags(dlbls, show_val="1", show_pct="0")
    _fix_val_axis(root, smart_max(vals))
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def process_chart_stacked(chart_xml: bytes, years: list, media_matrix: pd.DataFrame) -> bytes:
    root = etree.fromstring(chart_xml)
    bar_chart = root.find(f".//{ctag('barChart')}")
    if bar_chart is None:
        return chart_xml
    media_order = [m for m in ["AF","PR","RD","TV","CN"]
                   if m in media_matrix.columns and media_matrix[m].sum() > 0]
    existing_sers = bar_chart.findall(ctag("ser"))
    ref_ser = existing_sers[0] if existing_sers else None
    for s in existing_sers:
        bar_chart.remove(s)
    _ib = bar_chart.find(ctag("dLbls"))
    insert_before = _ib if _ib is not None else bar_chart.find(ctag("axId"))
    extra_colors = {"TV": "2563EB", "CN": "70AD47"}
    for idx, m in enumerate(media_order):
        orig_idx = ["AF","PR","RD","TV","CN"].index(m)
        if orig_idx < len(existing_sers):
            new_ser = deepcopy(existing_sers[orig_idx])
        elif ref_ser is not None:
            new_ser = deepcopy(ref_ser)
            if m in extra_colors:
                _set_series_color(new_ser, extra_colors[m])
        else:
            new_ser = etree.SubElement(bar_chart, ctag("ser"))
        idx_el = new_ser.find(ctag("idx"))
        if idx_el is None: idx_el = etree.SubElement(new_ser, ctag("idx"))
        idx_el.set("val", str(idx))
        order_el = new_ser.find(ctag("order"))
        if order_el is None: order_el = etree.SubElement(new_ser, ctag("order"))
        order_el.set("val", str(idx))
        _set_series_name(new_ser, m)
        vals = [media_matrix.loc[y, m] if y in media_matrix.index else 0 for y in years]
        vals_clean = [v if v and v > 0 else None for v in vals]
        _rebuild_cache(new_ser, [int(y) for y in years], vals_clean, divide_by=1e6)
        dlbls = new_ser.find(ctag("dLbls"))
        if dlbls is None: dlbls = etree.SubElement(new_ser, ctag("dLbls"))
        for child in list(dlbls): dlbls.remove(child)
        nf = etree.SubElement(dlbls, ctag("numFmt"))
        nf.set("formatCode", "0%"); nf.set("sourceLinked", "0")
        for tag, val in [("showLegendKey","0"),("showVal","0"),("showCatName","0"),
                         ("showSerName","0"),("showPercent","1"),("showBubbleSize","0")]:
            etree.SubElement(dlbls, ctag(tag)).set("val", val)
        if insert_before is not None:
            bar_chart.insert(list(bar_chart).index(insert_before), new_ser)
        else:
            bar_chart.append(new_ser)
    gdlbls = bar_chart.find(ctag("dLbls"))
    if gdlbls is None: gdlbls = etree.SubElement(bar_chart, ctag("dLbls"))
    for child in list(gdlbls): gdlbls.remove(child)
    etree.SubElement(gdlbls, ctag("dLblPos")).set("val", "ctr")
    nf = etree.SubElement(gdlbls, ctag("numFmt"))
    nf.set("formatCode", "0%"); nf.set("sourceLinked", "0")
    for tag, val in [("showLegendKey","0"),("showVal","0"),("showCatName","0"),
                     ("showSerName","0"),("showPercent","1"),("showBubbleSize","0")]:
        etree.SubElement(gdlbls, ctag(tag)).set("val", val)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def process_chart_seasonality(chart_xml: bytes, years: list, seas: pd.DataFrame) -> bytes:
    root = etree.fromstring(chart_xml)
    line_chart = root.find(f".//{ctag('lineChart')}")
    if line_chart is None:
        return chart_xml
    existing = line_chart.findall(ctag("ser"))
    for s in existing: line_chart.remove(s)
    _ib2 = line_chart.find(ctag("dLbls"))
    insert_before = _ib2 if _ib2 is not None else line_chart.find(ctag("axId"))
    ref_ser = existing[1] if len(existing) > 1 else (existing[0] if existing else None)
    all_vals = []
    for i, y in enumerate(years):
        orig_idx = i + 1
        if orig_idx < len(existing):
            new_ser = deepcopy(existing[orig_idx])
        elif ref_ser is not None:
            new_ser = deepcopy(ref_ser)
        else:
            new_ser = etree.SubElement(line_chart, ctag("ser"))
        idx_el = new_ser.find(ctag("idx"))
        if idx_el is None: idx_el = etree.SubElement(new_ser, ctag("idx"))
        idx_el.set("val", str(i))
        order_el = new_ser.find(ctag("order"))
        if order_el is None: order_el = etree.SubElement(new_ser, ctag("order"))
        order_el.set("val", str(i))
        month_vals = [seas.loc[y, m] if y in seas.index and m in seas.columns else 0
                      for m in range(1, 13)]
        all_vals.extend(month_vals)
        _rebuild_cache(new_ser, MONTHS_FR, month_vals, divide_by=1e6)
        _set_series_name(new_ser, str(int(y)))
        _build_peak_dlbls(new_ser, month_vals, top_n=3)
        if insert_before is not None:
            line_chart.insert(list(line_chart).index(insert_before), new_ser)
        else:
            line_chart.append(new_ser)
    _fix_val_axis(root, smart_max(all_vals))
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def process_chart_annonceurs(chart_xml: bytes, cats: list, vals: list, year_label: str) -> bytes:
    root = etree.fromstring(chart_xml)
    sers = root.findall(f".//{ctag('ser')}")
    if sers:
        cats = cats[:TOP_N_ANNONCEURS]; vals = vals[:TOP_N_ANNONCEURS]
        _rebuild_cache(sers[0], cats, vals, divide_by=1e6)
        _set_series_name(sers[0], year_label)
        dlbls = root.find(f".//{ctag('dLbls')}")
        if dlbls is not None:
            _set_num_fmt(dlbls, FMT_MILLIONS)
            _set_show_flags(dlbls, show_val="1", show_pct="0")
    _fix_val_axis(root, smart_max(vals))
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def process_chart_top_ann(chart_xml: bytes, cats: list, vals: list, label: str) -> bytes:
    root = etree.fromstring(chart_xml)
    sers = root.findall(f".//{ctag('ser')}")
    if sers:
        cats = cats[:TOP_N_ANNONCEURS]; vals = vals[:TOP_N_ANNONCEURS]
        _rebuild_cache(sers[0], cats, vals, divide_by=1e6)
        _set_series_name(sers[0], label)
        dlbls = root.find(f".//{ctag('dLbls')}")
        if dlbls is not None:
            _set_num_fmt(dlbls, FMT_MILLIONS)
            _set_show_flags(dlbls, show_val="1", show_pct="0")
    _fix_val_axis(root, smart_max(vals))
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def process_chart_pie(chart_xml: bytes, cats: list, vals: list, label: str) -> bytes:
    root = etree.fromstring(chart_xml)
    sers = root.findall(f".//{ctag('ser')}")
    if sers:
        _rebuild_cache(sers[0], cats, vals, divide_by=1e6)
        _set_series_name(sers[0], label)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


# ═══════════════════════════════════════════════════════════════
# 4. HELPERS XML — SLIDES & TEXTES
# ═══════════════════════════════════════════════════════════════
def _replace_tf_in_xml(txBody, new_text: str):
    pns = ANS
    def qtag(n): return f"{{{pns}}}{n}"
    font_sz = font_b = None
    orig_rpr = None
    for p in txBody.findall(qtag("p")):
        for r in p.findall(qtag("r")):
            rpr = r.find(qtag("rPr"))
            if rpr is not None:
                font_sz = rpr.get("sz")
                font_b  = rpr.get("b")
                orig_rpr = deepcopy(rpr)
                break
        if orig_rpr is not None: break
    first_pPr = None
    paras = txBody.findall(qtag("p"))
    if paras:
        first_pPr = paras[0].find(qtag("pPr"))
    for p in list(txBody.findall(qtag("p"))): txBody.remove(p)
    for line in new_text.split("\n"):
        p_el = etree.SubElement(txBody, qtag("p"))
        if first_pPr is not None: p_el.insert(0, deepcopy(first_pPr))
        if not line.strip():
            end = etree.SubElement(p_el, qtag("endParaRPr"))
            end.set("lang", "fr-FR"); end.set("dirty", "0")
            continue
        r_el = etree.SubElement(p_el, qtag("r"))
        if orig_rpr is not None:
            r_el.append(deepcopy(orig_rpr))
        else:
            rpr = etree.SubElement(r_el, qtag("rPr"))
            rpr.set("lang", "fr-FR"); rpr.set("dirty", "0")
            if font_sz: rpr.set("sz", font_sz)
            if font_b:  rpr.set("b", font_b)
        t_el = etree.SubElement(r_el, qtag("t"))
        t_el.text = line


def update_slide_texts(slide_xml: bytes, updates: dict) -> bytes:
    root = etree.fromstring(slide_xml)
    for sp in root.findall(f".//{ptag('sp')}"):
        nv = sp.find(f".//{ptag('cNvPr')}")
        if nv is None: continue
        name = nv.get("name", "")
        if name not in updates: continue
        txBody = sp.find(f".//{ptag('txBody')}")
        if txBody is None:
            txBody = sp.find(f".//{atag('txBody')}")
        if txBody is None: continue
        _replace_tf_in_xml(txBody, updates[name])
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


# ═══════════════════════════════════════════════════════════════
# 5. HELPERS EXCEL EMBARQUÉS
# ═══════════════════════════════════════════════════════════════
def _make_clean_workbook(emb_bytes: bytes) -> "openpyxl.Workbook":
    """
    Crée un workbook propre en partant du template mais avec une feuille entièrement
    réinitialisée — supprime toutes les cellules existantes pour éviter les cellules
    fantômes qui corrompent les charts quand PowerPoint recalcule.
    """
    import openpyxl as _xl
    from openpyxl import Workbook
    wb_orig = _xl.load_workbook(io.BytesIO(emb_bytes))
    sheet_name = wb_orig.active.title  # conserver le nom de feuille original ("Sheet1")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Désactiver fullCalcOnLoad : empêche PowerPoint de recalculer depuis les formules
    # et d'écraser le cache avec les données du template
    wb.calculation.fullCalcOnLoad = False
    wb.calculation.calcOnSave = False

    return wb, ws


def _xlsx_simple(emb_bytes: bytes, cats: list, vals: list, col_header: str = "Valeur") -> bytes:
    """Workbook 2 colonnes : catégories | valeurs — from scratch, sans cellules fantômes."""
    wb, ws = _make_clean_workbook(emb_bytes)
    ws.cell(1, 1).value = ""
    ws.cell(1, 2).value = col_header
    for i, (cat, val) in enumerate(zip(cats, vals)):
        ws.cell(i + 2, 1).value = cat
        ws.cell(i + 2, 2).value = round(float(val), 2) if val is not None else 0
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def _xlsx_multi(emb_bytes: bytes, cats: list, series_dict: dict) -> bytes:
    """Workbook multi-séries (saisonnalité) — from scratch, sans cellules fantômes."""
    wb, ws = _make_clean_workbook(emb_bytes)
    ws.cell(1, 1).value = ""
    for j, name in enumerate(series_dict.keys()):
        ws.cell(1, j + 2).value = str(name)
    for i, cat in enumerate(cats):
        ws.cell(i + 2, 1).value = cat
        for j, vals in enumerate(series_dict.values()):
            v = vals[i] if i < len(vals) else 0
            ws.cell(i + 2, j + 2).value = round(float(v), 2) if v is not None else 0
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def _xlsx_stacked(emb_bytes: bytes, cats: list, series_dict: dict) -> bytes:
    """Workbook stacked bar (mix médias) — from scratch, sans cellules fantômes."""
    wb, ws = _make_clean_workbook(emb_bytes)
    ws.cell(1, 1).value = "Colonne1"
    for j, name in enumerate(series_dict.keys()):
        ws.cell(1, j + 2).value = name
    for i, cat in enumerate(cats):
        ws.cell(i + 2, 1).value = cat
        for j, (name, vals) in enumerate(series_dict.items()):
            v = vals[i] if i < len(vals) else 0
            ws.cell(i + 2, j + 2).value = round(float(v or 0), 2)
    out = io.BytesIO(); wb.save(out); return out.getvalue()


# ═══════════════════════════════════════════════════════════════
# 6. INJECTION PRINCIPALE — PPTInjector
# ═══════════════════════════════════════════════════════════════
class PPTInjector:
    """
    Génère un PPT dynamique:
    - Slides 1, 2, 3 : toujours présentes
    - Slides médias  : une slide par média présent dans la data
    - Ordre de sortie: AF → TV → RD → PR → CN
    """

    def __init__(self, template_path: str):
        self.template_path = template_path
        with open(template_path, "rb") as f:
            self.template_raw = f.read()

    def generate(self, calc: MediaCalculator, comments: dict,
                 secteur: str, sous_secteur: str) -> bytes:

        years     = calc.years
        year_last = calc.year_last
        totals    = calc.total_by_year()
        mm        = calc.total_by_year_media()
        seas      = calc.seasonality_by_year()
        label     = sous_secteur or secteur
        yrange    = f"{years[0]} – {year_last}" if years else ""
        medias    = calc.medias_present  # déjà dans l'ordre MEDIA_ORDER

        # ── Lire le template ZIP ──────────────────────────────────────
        original = {}
        with zipfile.ZipFile(io.BytesIO(self.template_raw), "r") as zin:
            for item in zin.infolist():
                original[item.filename] = zin.read(item.filename)

        # ── Helper: récupérer l'embedding d'un chart ─────────────────
        def _emb(chart_id):
            rp = f"ppt/charts/_rels/chart{chart_id}.xml.rels"
            if rp not in original: return None, None
            rr = etree.fromstring(original[rp])
            for rel in rr:
                t = rel.get("Target", "")
                if "embedding" in t:
                    name = t.split("/")[-1]
                    path = f"ppt/embeddings/{name}"
                    return path, original.get(path)
            return None, None

        # ── Helper: extraire rId → chart_num depuis rels d'une slide ─
        def _slide_chart_map(slide_num):
            rp = f"ppt/slides/_rels/slide{slide_num}.xml.rels"
            if rp not in original: return {}
            rels = etree.fromstring(original[rp])
            result = {}
            for rel in rels:
                if "chart" in rel.get("Target", ""):
                    cname = rel.get("Target", "").split("/")[-1]
                    cnum = int(cname.replace("chart", "").replace(".xml", ""))
                    result[rel.get("Id")] = cnum
            return result

        # ────────────────────────────────────────────────────────────
        # SLIDES 1-3 : toujours générées
        # ────────────────────────────────────────────────────────────
        chart_updates = {}
        excel_updates = {}

        # Slide 2 — Overview
        ch1 = process_chart_annual(
            original["ppt/charts/chart1.xml"],
            [int(y) for y in years], [totals.get(y, 0) for y in years], "Total"
        )
        ch2 = process_chart_seasonality(original["ppt/charts/chart2.xml"], years, seas)
        ch3 = process_chart_stacked(original["ppt/charts/chart3.xml"], years, mm)
        chart_updates.update({"ppt/charts/chart1.xml": ch1,
                               "ppt/charts/chart2.xml": ch2,
                               "ppt/charts/chart3.xml": ch3})

        p, b = _emb(1)
        if p and b:
            excel_updates[p] = _xlsx_simple(b, [int(y) for y in years],
                                             [totals.get(y, 0) for y in years], "Total")
        p, b = _emb(2)
        if p and b:
            month_data = {str(int(y)): [seas.loc[y, m] if y in seas.index and m in seas.columns else 0
                                         for m in range(1, 13)] for y in years}
            excel_updates[p] = _xlsx_multi(b, MONTHS_FR, month_data)
        p, b = _emb(3)
        if p and b:
            media_present_list = [m for m in ["AF","PR","RD","TV","CN"]
                                   if m in mm.columns and mm[m].sum() > 0]
            stacked_data = {m: [mm.loc[y, m] if y in mm.index else 0 for y in years]
                            for m in media_present_list}
            excel_updates[p] = _xlsx_stacked(b, [int(y) for y in years], stacked_data)

        # Slide 3 — Annonceurs
        years_ann = years[-3:] if len(years) >= 3 else years
        for chart_id, y in zip([4, 5, 6], list(years_ann) + [None] * (3 - len(years_ann))):
            if y is None: continue
            top = calc.top_annonceurs_by_year(y, n=TOP_N_ANNONCEURS)
            xml = process_chart_annonceurs(
                original[f"ppt/charts/chart{chart_id}.xml"],
                list(top.index), list(top.values), str(int(y))
            )
            chart_updates[f"ppt/charts/chart{chart_id}.xml"] = xml
            p, b = _emb(chart_id)
            if p and b:
                excel_updates[p] = _xlsx_simple(b, list(top.index), list(top.values), str(int(y)))

        # Texts slides 1-3
        slide_texts = {
            "ppt/slides/slide1.xml": {
                "Title 1": f"Media Review\n{label}\n| {yrange}",
            },
            "ppt/slides/slide2.xml": {
                "Text 0":    f"Investissements média — {label}",
                "Text 1":    f"{yrange} | Millions MAD | Source : Imperium",
                "TextBox 15": (comments.get("slide2_headline", "") + "\n\n" +
                               comments.get("slide2_global", "")),
            },
            "ppt/slides/slide3.xml": {
                "TextBox 1":    f"Investissement média par annonceur — {label}",
                "TextBox 2":    f"Classement annonceurs | Millions MAD | {' - '.join(str(int(y)) for y in years[-3:])} | Source : Imperium",
                "ZoneTexte 13": comments.get("slide3_annonceurs", ""),
            },
        }

        # ────────────────────────────────────────────────────────────
        # SLIDES MÉDIAS — dynamiques
        # ────────────────────────────────────────────────────────────
        # Mapping template_slide_num → données réelles pour chaque média
        # On va construire le PPT final en:
        # 1. Gardant slides 1, 2, 3
        # 2. Pour chaque média présent: copier la slide template correspondante
        # 3. Supprimer toutes les slides médias du template (4-8)
        # 4. Assembler dans le bon ordre

        # D'abord, préparer les charts pour chaque média présent
        media_chart_data = {}  # code → {annual_xml, pie_xml, topann_xml, emb_annual, emb_pie, emb_topann}

        for template_slide_num, (code, rid_annual, rid_pie, rid_topann) in TEMPLATE_SLIDE_MAP.items():
            if code not in medias:
                continue

            chart_map = _slide_chart_map(template_slide_num)
            cid_annual  = chart_map.get(rid_annual)
            cid_pie     = chart_map.get(rid_pie)
            cid_topann  = chart_map.get(rid_topann)

            mt = calc.total_focus_by_year(code)
            ta = calc.top_ann_focus_last(code)
            ts = calc.split_support_last(code)
            media_label_full = MEDIA_LABELS.get(code, code)
            media_short = MEDIA_SHORT.get(code, code)

            annual_xml = process_chart_annual(
                original[f"ppt/charts/chart{cid_annual}.xml"],
                [int(y) for y in years], [mt.get(y, 0) for y in years], media_label_full
            )
            pie_xml = process_chart_pie(
                original[f"ppt/charts/chart{cid_pie}.xml"],
                list(ts.index), list(ts.values), media_short
            )
            topann_xml = process_chart_top_ann(
                original[f"ppt/charts/chart{cid_topann}.xml"],
                list(ta.index), list(ta.values), f"Top {media_short} {int(year_last)}"
            )

            # Excel embeddings
            p_ann, b_ann = _emb(cid_annual)
            p_pie, b_pie = _emb(cid_pie)
            p_top, b_top = _emb(cid_topann)
            emb_annual = _xlsx_simple(b_ann, [int(y) for y in years],
                                       [mt.get(y, 0) for y in years], media_short) if b_ann else None
            emb_pie    = _xlsx_simple(b_pie, list(ts.index), list(ts.values), media_short) if b_pie else None
            emb_topann = _xlsx_simple(b_top, list(ta.index), list(ta.values), media_short) if b_top else None

            media_chart_data[code] = {
                "template_slide": template_slide_num,
                "cid_annual": cid_annual, "cid_pie": cid_pie, "cid_topann": cid_topann,
                "annual_xml": annual_xml, "pie_xml": pie_xml, "topann_xml": topann_xml,
                "p_annual": p_ann, "emb_annual": emb_annual,
                "p_pie": p_pie,    "emb_pie": emb_pie,
                "p_top": p_top,    "emb_topann": emb_topann,
                "comment_shape": MEDIA_SLIDE_SHAPES[template_slide_num]["comment"],
            }

            # Mettre à jour les charts
            chart_updates[f"ppt/charts/chart{cid_annual}.xml"] = annual_xml
            chart_updates[f"ppt/charts/chart{cid_pie}.xml"]    = pie_xml
            chart_updates[f"ppt/charts/chart{cid_topann}.xml"] = topann_xml
            if p_ann and emb_annual: excel_updates[p_ann] = emb_annual
            if p_pie and emb_pie:   excel_updates[p_pie] = emb_pie
            if p_top and emb_topann: excel_updates[p_top] = emb_topann

        # ────────────────────────────────────────────────────────────
        # CONSTRUCTION DU ZIP FINAL
        # ────────────────────────────────────────────────────────────

        # Slides à garder : 1, 2, 3 + slides médias (dans l'ordre MEDIA_ORDER)
        # Slides à supprimer du template : toutes les slides 4-8 (remplacées dynamiquement)

        # Identifier quelles slides template on garde et quelles on supprime
        template_media_slides = {4, 5, 6, 7, 8}
        slides_to_skip = set()

        # Slides médias absentes de la data → supprimer slide + charts + embeddings
        for tslide, (code, *_) in TEMPLATE_SLIDE_MAP.items():
            if code not in medias:
                slides_to_skip.add(f"ppt/slides/slide{tslide}.xml")
                slides_to_skip.add(f"ppt/slides/_rels/slide{tslide}.xml.rels")
                # Supprimer aussi les charts et leurs embeddings (orphelins sinon)
                chart_map = _slide_chart_map(tslide)
                for cnum in chart_map.values():
                    slides_to_skip.add(f"ppt/charts/chart{cnum}.xml")
                    slides_to_skip.add(f"ppt/charts/_rels/chart{cnum}.xml.rels")
                    # Trouver l'embedding lié à ce chart
                    rels_path = f"ppt/charts/_rels/chart{cnum}.xml.rels"
                    if rels_path in original:
                        rels_root = etree.fromstring(original[rels_path])
                        for rel in rels_root:
                            if "embedding" in rel.get("Target", ""):
                                emb_name = rel.get("Target", "").split("/")[-1]
                                slides_to_skip.add(f"ppt/embeddings/{emb_name}")

        # Les slides médias présentes gardent leurs numéros de template pour l'instant
        # On génère le texte pour chaque
        for code, data in media_chart_data.items():
            tslide = data["template_slide"]
            sup_lbl = MEDIA_SUP_LABEL.get(code, "support")
            comment_key = f"slide_{code.lower()}"
            comment_shape = data["comment_shape"]

            texts = {
                "Text 0":    f"Investissements média {MEDIA_LABELS.get(code,code)} — {label}",
                "TextBox 3": f"FY {yrange} | Millions MAD | Source : Imperium",
                "TextBox 8": f"Investissements {MEDIA_SHORT.get(code,code)}",
                "TextBox 9": f"Répartition {int(year_last)} par {sup_lbl}",
                "TextBox 10": f"Top annonceurs {MEDIA_SHORT.get(code,code)} (FY {int(year_last)})",
                comment_shape: comments.get(comment_key, ""),
            }
            slide_texts[f"ppt/slides/slide{tslide}.xml"] = texts

        # Calculer le nouvel ordre des slides dans presentation.xml
        # Ordre final: slide1, slide2, slide3, [slides médias présentes dans ordre MEDIA_ORDER]
        media_slide_order = []
        for code in MEDIA_ORDER:
            if code in media_chart_data:
                tslide = media_chart_data[code]["template_slide"]
                media_slide_order.append(tslide)

        final_slide_order = [1, 2, 3] + media_slide_order

        # Lire presentation.xml.rels pour avoir le mapping rId → slide
        prs_rels_root = etree.fromstring(original["ppt/_rels/presentation.xml.rels"])
        rid_to_slide = {}
        slide_to_rid = {}
        for rel in prs_rels_root:
            target = rel.get("Target", "")
            if "slide" in target and "layout" not in target and "Master" not in target:
                sname = target.split("/")[-1]
                snum = int(sname.replace("slide", "").replace(".xml", ""))
                rid_to_slide[rel.get("Id")] = snum
                slide_to_rid[snum] = rel.get("Id")

        # ── Écrire le ZIP final ───────────────────────────────────────
        out_zip = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(self.template_raw), "r") as zin:
            with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    fname = item.filename

                    # Ignorer les slides médias absentes
                    if fname in slides_to_skip:
                        continue

                    data = zin.read(fname)

                    if fname in chart_updates:
                        data = chart_updates[fname]

                    elif fname in slide_texts:
                        try:
                            data = update_slide_texts(data, slide_texts[fname])
                        except Exception:
                            pass

                    elif fname == "ppt/presentation.xml":
                        data = self._reorder_slides_in_presentation(
                            data, final_slide_order, slide_to_rid)

                    elif fname == "ppt/_rels/presentation.xml.rels":
                        data = self._reorder_rels(data, slides_to_skip, slide_to_rid)

                    elif fname == "[Content_Types].xml":
                        data = self._fix_content_types(data, slides_to_skip)

                    elif fname in excel_updates:
                        data = excel_updates[fname]

                    zout.writestr(item, data)

        return out_zip.getvalue()

    def _reorder_slides_in_presentation(self, prs_xml: bytes,
                                         final_order: list, slide_to_rid: dict) -> bytes:
        """Réordonne les sldId dans presentation.xml selon final_order."""
        root = etree.fromstring(prs_xml)
        PRS_NS = PNS
        R_NS   = RNS

        sld_id_lst = root.find(f"{{{PRS_NS}}}sldIdLst")
        if sld_id_lst is None:
            return prs_xml

        # Construire mapping id_attr → sldId element
        existing = {sld.get(f"{{{R_NS}}}id"): sld for sld in sld_id_lst}

        # Vider et remettre dans le bon ordre
        for sld in list(sld_id_lst):
            sld_id_lst.remove(sld)

        for snum in final_order:
            rid = slide_to_rid.get(snum)
            if rid and rid in existing:
                sld_id_lst.append(existing[rid])

        return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

    def _reorder_rels(self, rels_xml: bytes, slides_to_skip: set,
                       slide_to_rid: dict) -> bytes:
        """Supprime les rels des slides à exclure."""
        root = etree.fromstring(rels_xml)
        # Construire set de slides à supprimer
        skip_rids = set()
        rid_to_target = {}
        for rel in root:
            target = rel.get("Target", "")
            full = f"ppt/{target}" if not target.startswith("ppt") else target
            rid_to_target[rel.get("Id", "")] = full
            if full in slides_to_skip:
                skip_rids.add(rel.get("Id", ""))
        for rel in list(root):
            if rel.get("Id", "") in skip_rids:
                root.remove(rel)
        return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

    def _fix_content_types(self, ct_xml: bytes, slides_to_skip: set) -> bytes:
        """Supprime les entrées Content-Types des slides exclues."""
        root = etree.fromstring(ct_xml)
        CT_NS = CTNS
        for override in list(root.findall(f"{{{CT_NS}}}Override")):
            part = override.get("PartName", "").lstrip("/")
            if part in slides_to_skip:
                root.remove(override)
        return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
